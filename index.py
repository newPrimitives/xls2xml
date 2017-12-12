from openpyxl import load_workbook # excell maniuplation library 
import xml.etree.cElementTree as ET # xml tree generator 
import glob # file filter 
import datetime # used for naming files 
import os # used to create folder for storing xml

# Indent XML into a beutified format 
def indent(elem, level=0):
  i = "\n" + level*"  "
  if len(elem):
    if not elem.text or not elem.text.strip():
      elem.text = i + "  "
    if not elem.tail or not elem.tail.strip():
      elem.tail = i
    for elem in elem:
      indent(elem, level+1)
    if not elem.tail or not elem.tail.strip():
      elem.tail = i
  else:
    if level and (not elem.tail or not elem.tail.strip()):
      elem.tail = i

# Load from excel and convert to XML 
def buildTree(sheet_name, header_row, content_row):
	workbook = ET.Element("Workbook")
	sheet = ET.SubElement(workbook, "Worksheet", name=sheet_name)

	header = ET.SubElement(sheet, "Row")
	for data in header_row:
		cell = ET.SubElement(header, "Cell")
		ET.SubElement(cell, "Data", type="String").text = str(data.value)

	row = ET.SubElement(sheet, "Row")
	for data in content_row:
		cell = ET.SubElement(row, "Cell")
		ET.SubElement(cell, "Data", type=type(data.value).__name__).text = str(data.value)

	indent(workbook)

	tree = ET.ElementTree(workbook)
	filename = "xml/" + str(datetime.datetime.now().date()) + '_' + str(datetime.datetime.now().time()).replace(':', '.') + '.xml'
  	tree.write(filename, xml_declaration=True, encoding='utf-8', method="xml")
 		
'''
main function, to run the script call python index.py
'''
if __name__ == "__main__":
	if not os.path.exists('xml'):
		os.makedirs('xml')

	types = ('*.xlsx', '*.xlsm', "*.xltx", "*.xltm")
	files_grabbed = []
	for files in types:
		files_grabbed.extend(glob.glob(files))

	for f in files_grabbed: 
		wb = load_workbook(f)
		for sheet in wb.worksheets:
			header = list(sheet.rows)[0]
	 		if sheet.max_row > 1:
	 			row_num = sheet.max_row
	 			for row_index in range(2, row_num + 1):
					buildTree(sheet.title, header, sheet[row_index])